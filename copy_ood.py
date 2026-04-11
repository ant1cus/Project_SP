import os
import re
import traceback
from pathlib import Path
from datetime import datetime
from speedcopy import copyfile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def write_error(text: str, color: str, err: list, cell, log) -> bool:
    log.error(re.sub('replace', str(cell.value), text))
    err.append(re.sub('replace', str(cell.value), text))
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    return True


def copy_files(incoming_data: dict, current_progress, now_doc, all_doc, line_doing, line_progress, progress_value,
               event, window_check, info_value) -> dict:
    color_cell = {
        # {get_column_letter(column + 1)}
        'zero': {'text': f"Количество снимков для sn replace равно 0", 'color': 'FF0000'},
        'more': {'text': f"Количество снимков для sn replace больше указанного", 'color': '5252FF'},
        'less': {'text': f"Количество снимков для sn replace меньше указанного", 'color': 'FFFF00'},
    }
    logging = incoming_data['logging']
    errors = []
    try:
        start_time = datetime.now()
        logging.info(f"время начала - {start_time}")
        name_finish_folder = incoming_data['name_gk'] if incoming_data['name_gk'] else 'Номер ГК'
        if Path(incoming_data['finish_path'], name_finish_folder).exists() is False:
            os.makedirs(Path(incoming_data['finish_path'], name_finish_folder))
        logging.info(f"Считываем файл выгрузки {Path(incoming_data['upload_file']).name}")
        df = pd.read_excel(Path(incoming_data['upload_file']), sheet_name=0, header=None, dtype=str)
        if incoming_data['name_set']:
            df.fillna(value={2: incoming_data['name_set']}, inplace=True)
        df[1] = df[1].astype(str)
        df[3] = df[3].astype(str)
        df['finish_folder'] = str(Path(incoming_data['finish_path'], name_finish_folder)) + '\\' + df[1] + '\\' + df[3]
        serial_nums = {}
        logging.info(f"загрузка excel - {datetime.now() - start_time}")
        # Для кол-ва копий можно будет потом расскоментировать
        snapshot_files = {}
        all_poss_files = 1
        for index, item in df.iloc[0].items():
            if re.findall(r'\d+', str(item)):
                if index == 'finish_folder':
                    continue
                # folder = int(str(item).partition('-')[0]) if re.findall(r'\d+-\d+', str(item)) else int(item)
                # snapshot = int(str(item).partition('-')[2]) if re.findall(r'\d+-\d+', str(item)) else 0
                folder = int(str(item).partition('-')[0])
                snapshot = int(str(item).partition('-')[2])
                full_path = dict(zip(
                    df[index].iloc[2:],
                    [f"{i}\\{folder}" for i in df['finish_folder'].iloc[2:].to_numpy().tolist()]
                ))
                full_find = dict(zip(
                    df[index].iloc[2:],
                    [{'coordinate': val, 'more_snapshot': False, 'snapshot': [False for _ in range(0, snapshot)]}
                     for val in {i: [index, i] for i in df.index.to_list()[2:]}.values()]
                ))
                serial_nums.update(full_path)
                # serial_snapshot.update(full_snapshot)
                snapshot_files.update(full_find)
                all_poss_files += len(full_path)*int(snapshot)
        percent = 100 / all_poss_files
        logging.info(f"считывание excel - {datetime.now() - start_time}")
        for file in Path(incoming_data['start_path']).rglob('*.*'):
            if window_check.stop_threading:
                return {'status': 'cancel', 'trace': '', 'text': ''}
            file_sn = file.stem.partition('_')[2].partition('_')[2].partition('_')[0]
            try:
                if file_sn not in serial_nums:
                    continue
                if not Path(serial_nums[file_sn], 'photo').exists():
                    os.makedirs(Path(serial_nums[file_sn], 'photo'))
                if Path(serial_nums[file_sn], 'photo', file.name).exists() is False:
                    line_doing.emit(f'Копируем файл {file.name} ({now_doc} из {all_doc})')
                    copyfile(str(file), str(Path(serial_nums[file_sn], 'photo', file.name)))
            except BaseException as ex:
                logging.error(f"Копирование файла {file.name} не завершено из-за ошибки - {ex}")
                logging.error(traceback.format_exc())
                errors.append(f"Файл {file.name} не скопирован из-за непредвиденной ошибки")
            try:
                snapshot = int(file.stem.partition('_')[2].partition('_')[0])
                if snapshot <= len(snapshot_files[file_sn]['snapshot']):
                    snapshot_files[file_sn]['snapshot'][snapshot - 1] = True
                else:
                    logging.warning(f"У файла {file.name} ракурс больше, чем указано в файле")
                    snapshot_files[file_sn]['more_snapshot'] = True
            except ValueError:
                logging.error(f"Количество снимков в файле {file.name} не конвертируется в целое число")
                errors.append(f"Количество снимков в файле {file.name} не конвертируется в целое число")
            now_doc += 1
            current_progress += percent
            line_progress.emit(f'Выполнено {int(current_progress)} %')
            progress_value.emit(int(current_progress))
        # Проверки
        logging.info(f"скопировали все файлы - {datetime.now() - start_time}")
        xlsx_path = Path(incoming_data['upload_file'])
        wb = load_workbook(str(xlsx_path))
        ws = wb.active
        write_xlsx = False
        for sn in snapshot_files:
            cell = f"{get_column_letter(snapshot_files[sn]['coordinate'][0] + 1)}" \
                   f"{snapshot_files[sn]['coordinate'][1] + 1}"
            if snapshot_files[sn]['more_snapshot']:
                write_xlsx = write_error(color_cell['more']['text'], color_cell['more']['color'], errors, ws[cell],
                                         logging)
            if all(snapshot_files[sn]['snapshot']):
                continue
            if not any(snapshot_files[sn]['snapshot']):
                write_xlsx = write_error(color_cell['zero']['text'], color_cell['zero']['color'], errors, ws[cell],
                                         logging)
            elif not all(snapshot_files[sn]['snapshot']):
                write_xlsx = write_error(color_cell['less']['text'], color_cell['less']['color'], errors, ws[cell],
                                         logging)
        logging.info(f"пробежались по всем ошибкам - {datetime.now() - start_time}")
        if write_xlsx:
            while True:
                try:
                    wb.save(str(xlsx_path))
                    break
                except PermissionError:
                    info_value.emit('Вопрос?',
                                    f"Файл {xlsx_path.name} открыт. Для сохранения ошибок в файл закройте "
                                    f"его и нажмите «Да». Если нажать «Нет» ошибки не будут посдвечены!",
                                    "")
                    event.clear()
                    event.wait()
                if window_check.answer is False:
                    break
        logging.info(f"конец программы - {datetime.now() - start_time}")
        return {'status': 'success', 'text': '', 'trace': ''}
    except BaseException as ex:
        return {'status': 'error', 'text': ex, 'trace': traceback.format_exc()}
