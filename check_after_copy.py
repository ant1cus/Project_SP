import os
import re
import traceback
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def write_error(text: str, color: str, err: list, cell, log) -> bool:
    log.error(re.sub('replace', str(cell.value), text))
    err.append(re.sub('replace', str(cell.value), text))
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    return True


def check_after_copy(incoming_data: dict, current_progress, now_doc, all_doc, line_doing, line_progress,
					 progress_value, event, window_check, info_value) -> dict:
	color_cell = {
		# {get_column_letter(column + 1)}
		'zero': {'text': f"Количество снимков для sn replace равно 0", 'color': 'FF0000'},
		'more': {'text': f"Количество снимков для sn replace больше указанного", 'color': '5252FF'},
		'less': {'text': f"Количество снимков для sn replace меньше указанного", 'color': 'FFFF00'},
	}
	logging = incoming_data['logging']
	errors = []
	try:
		df = pd.read_excel(Path(incoming_data['upload_file']), sheet_name=0, header=None, dtype=str)
		df[1] = df[1].astype(str)
		df[3] = df[3].astype(str)
		df['finish_folder'] = str(Path(incoming_data['start_path'])) + '\\' + df[3]
		serial_nums = {}
		snapshot_files = {}
		all_poss_files = 1
		for index, item in df.iloc[0].items():
			if re.findall(r'\d+', str(item)):
				if index == 'finish_folder':
					continue
				folder = int(str(item).partition('-')[0])
				snapshot = int(str(item).partition('-')[2])
				full_path = dict(zip(
					df[index].iloc[2:],
					[f"{i}\\{folder}" for i in df['finish_folder'].iloc[2:].to_numpy().tolist()]
				))
				full_find = dict(zip(
					df[index].iloc[2:],
					[{'coordinate': val, 'more_snapshot': False, 'snapshot': [False for _ in range(0, snapshot)]}
					 for val in {i: [index, i] for i in df.index.to_list()[2:]}.values()]
				))
				serial_nums.update(full_path)
				snapshot_files.update(full_find)
				all_poss_files += len(full_path) * int(snapshot)
		percent = 10
		all_doc = all_poss_files
		for file in Path(incoming_data['start_path']).rglob('*.*'):
			if window_check.stop_threading:
				return {'status': 'cancel', 'trace': '', 'text': ''}
			file_sn = file.stem.partition('_')[2].partition('_')[2].partition('_')[0]
			try:
				if file_sn not in serial_nums:
					continue
				if len(file.name.split('_')) >= 4:
					errors.append(f"Файл {file.name} перенесен с номером коробки")
				if not Path(serial_nums[file_sn], 'photo').exists() and\
						f"Пути {Path(serial_nums[file_sn], 'photo')} не существует" not in errors:
					errors.append(f"Пути {Path(serial_nums[file_sn], 'photo')} не существует")
				snapshot = int(file.stem.partition('_')[2].partition('_')[0])
				if snapshot <= len(snapshot_files[file_sn]['snapshot']):
					snapshot_files[file_sn]['snapshot'][snapshot - 1] = True
				else:
					logging.warning(f"У файла {file.name} ракурс больше, чем указано в файле")
					snapshot_files[file_sn]['more_snapshot'] = True
				now_doc += 1
				if current_progress == 100:
					current_progress = 0
					progress_value.emit(int(current_progress))
				current_progress += percent
				progress_value.emit(int(current_progress))
			except BaseException as ex:
				logging.error(f"Проверка файла {file.name} не завершена из-за ошибки: {ex}")
				errors.append(f"Проверка файла {file.name} не завершена из-за непредвиденной ошибки")
		# Проверки
		xlsx_path = Path(incoming_data['upload_file'])
		wb = load_workbook(str(xlsx_path))
		ws = wb.active
		write_xlsx = False
		for sn in snapshot_files:
			cell = f"{get_column_letter(snapshot_files[sn]['coordinate'][0] + 1)}" \
				   f"{snapshot_files[sn]['coordinate'][1] + 1}"
			if snapshot_files[sn]['more_snapshot']:
				write_xlsx = write_error(color_cell['more']['text'], color_cell['more']['color'], errors, ws[cell],
										 logging)
			if all(snapshot_files[sn]['snapshot']):
				continue
			if not any(snapshot_files[sn]['snapshot']):
				write_xlsx = write_error(color_cell['zero']['text'], color_cell['zero']['color'], errors, ws[cell],
										 logging)
			elif not all(snapshot_files[sn]['snapshot']):
				write_xlsx = write_error(color_cell['less']['text'], color_cell['less']['color'], errors, ws[cell],
										 logging)
		if write_xlsx:
			while True:
				try:
					wb.save(str(xlsx_path))
					break
				except PermissionError:
					info_value.emit('Вопрос?',
									f"Файл {xlsx_path.name} открыт. Для сохранения ошибок в файл закройте "
									f"его и нажмите «Да». Если нажать «Нет» ошибки не будут выделены!",
									"")
					event.clear()
					event.wait()
				if window_check.answer is False:
					break
			os.startfile(xlsx_path)
		return {'status': 'success', 'text': '', 'trace': ''}
	except BaseException as ex:
		return {'status': 'error', 'text': ex, 'trace': traceback.format_exc()}

