import math
import os
import re
import traceback
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import PatternFill
import pandas as pd


def copy_from_asu_file(incoming_data: dict, current_progress: float, now_doc: int, all_doc: int, line_doing,
                       line_progress, progress_value, event, window_check, info_value) -> dict:
    logging = incoming_data['logging']
    try:
        # Определяем фотки в папках или отдельно
        # Для СП
        percent = incoming_data['percent']
        name_finish_folder = incoming_data['name_gk'] if incoming_data['name_gk'] else 'Номер ГК'
        try:
            os.mkdir(Path(incoming_data['path_finish_folder'], name_finish_folder))
        except FileExistsError:
            logging.info(f"Такая папка уже есть {Path(incoming_data['path_finish_folder'], name_finish_folder)}")
        logging.info('Считываем файлы АСУ и создаем структуру каталога')
        # Для ООД
        try:
            logging.info(f"Считываем файл выгрузки {Path(incoming_data['path_load_asu']).name}")
            df = pd.read_excel(Path(incoming_data['path_load_asu']), sheet_name=0, header=None)
            index_string = -100
            number_snapshot = {}
            index_snapshot = []
            folder_number = []
            for index, item in df.iloc[0].items():
                if re.findall(r'\d+', str(item)):
                    if re.findall(r'\d+-\d+', str(item)):
                        number_snapshot[index] = int(str(item).partition('-')[2])
                        df.iloc[0, index] = int(str(item).partition('-')[0])
                    else:
                        number_snapshot[index] = 0
                        df.iloc[0, index] = int(item)
                    folder_number.append(df.iloc[0, index])
                    index_snapshot.append(index)
                    continue
                if re.findall(r'[A-zА-я]+', str(item)) or math.isnan(item):
                    df.iloc[0, index] = index_string
                    index_string += 1
            df.sort_values(0, axis=1, inplace=True)
            df = df.drop(labels=[1], axis=0)
            if incoming_data['name_set']:
                df.fillna(value={2: incoming_data['name_set']}, inplace=True)
            columns_name = ['name', 'snapshot', 'sn', 'start_path', 'parent_path', 'parent_name', 'sn_set',
                            'folder_number', 'name_set', 'copy_files', 'rename_file']
            documents = pd.DataFrame(columns=columns_name)
            errors = []
            line_doing.emit(f"Считываем снимки")
            rename_col = 3 if incoming_data['executor'] == 'sp' else 2
            for file in Path(incoming_data['path_material_sp']).rglob('*.*'):
                name_file = file.stem
                copy_file = True if 'info' in name_file.lower() or 'spk' in name_file.lower() else False
                sn_set = 0
                name_set = ''
                rename_file = ''
                if copy_file:
                    sn_info_file = file.name.partition('_')[2].partition('.')[0]
                    double_zero_sn = '00' + sn_info_file
                    index_info_file = []
                    for i in range(2, df.shape[1]):
                        exit_ = False
                        snap_val = df.loc[0, i]
                        df[i] = df[i].str.lower()
                        df.loc[0, i] = snap_val
                        for j in [sn_info_file, double_zero_sn]:
                            index_info_file = df.loc[df[i] == j.lower()].index.to_list()
                            if len(index_info_file) > 0:
                                rename_file = df.loc[index_info_file[0], rename_col]
                                exit_ = True
                                break
                        if exit_:
                            break
                    sn_set = 0 if len(index_info_file) == 0 else df.loc[index_info_file[0], 3]
                    name_set = '' if len(index_info_file) == 0 else df.loc[index_info_file[0], 2]
                documents = pd.concat([
                    documents,
                    pd.DataFrame({'name': [file.name], 'start_path': [file],
                                  'parent_path': [file.parent], 'parent_name': [file.parent.name],
                                  'snapshot': [name_file.partition('_')[2].partition('_')[0]],
                                  'sn': [name_file.partition('_')[2].partition('_')[2].partition('_')[0].lower()],
                                  'sn_set': [sn_set], 'folder_number': [0], 'name_set': [name_set],
                                  'copy_files': [copy_file], 'rename_file': [rename_file]
                                  })], ignore_index=True)
            logging.info(f"Проверяем на соответствие количества снимков")
            line_doing.emit("Проверяем на соответствие количества снимков")
            snapshot_df = df[index_snapshot]
            color_cell = {}
            color_column = {}
            index_for_snapshot = list(snapshot_df.index)
            for column in index_snapshot:
                number_snap = int(number_snapshot[column])
                if number_snap == 0:
                    errors.append(f"Количество снимков в столбце {get_column_letter(column + 1)} не указано")
                    color_column[column + 1] = '00FF00'
                for index, value in enumerate(snapshot_df[column].to_numpy().tolist()[1:]):
                    index_doc = documents.loc[documents['sn'] == str(value).lower()].index
                    if len(index_doc) == 0:
                        index_doc = documents.loc[documents['parent_name'] == value].index
                    documents.loc[index_doc, 'sn_set'] = df.loc[index_for_snapshot[index + 1], 3]
                    documents.loc[index_doc, 'folder_number'] = snapshot_df.loc[0, column]
                    documents.loc[index_doc, 'name_set'] = df.loc[index_for_snapshot[index + 1], 2]
                    if number_snap == 0:
                        continue
                    if len(index_doc) == 0:
                        errors.append(f"Количество снимков для sn {value} равно 0")
                        color_cell[f"{get_column_letter(column + 1)}{index + 3}"] = 'FF0000'
                    elif len(index_doc) > number_snap:
                        errors.append(f"Количество снимков для sn {value} больше указанного")
                        color_cell[f"{get_column_letter(column + 1)}{index + 3}"] = 'FFFF00'
                    elif len(index_doc) < number_snap:
                        errors.append(f"Количество снимков для sn {value} меньше указанного")
                        color_cell[f"{get_column_letter(column + 1)}{index + 3}"] = 'FFFF00'
                    else:
                        continue
            if len(color_column) == df.shape[1] - 4:
                errors = []
            else:
                wb = load_workbook(str(Path(incoming_data['path_load_asu'])))
                ws = wb.active
                for cell in color_cell:
                    ws[cell].fill = PatternFill(start_color=color_cell[cell], end_color=color_cell[cell],
                                                fill_type='solid')
                wb.save(str(Path(incoming_data['path_load_asu'])))
                for column in color_column:
                    for row in range(1, ws.max_row + 1):
                        ws.cell(row=row, column=column).fill = PatternFill(start_color=color_column[column],
                                                                           end_color=color_column[column],
                                                                           fill_type='solid')
                wb.save(str(Path(incoming_data['path_load_asu'])))
            if errors:
                info_value.emit('Вопрос?', '\n'.join(errors), 'Найдены несоответствия, для продолжения нажмите «Да»')
                event.clear()
                event.wait()
            if window_check.answer is False:
                return {'status': 'cancel', 'text': '', 'trace': ''}
            errors = []
        except BaseException as error:
            return {'status': 'error', 'text': error, 'trace': traceback.format_exc()}
        logging.info(f"Бежим по файлам, функция - {incoming_data['function']}")
        answer = incoming_data['function'](incoming_data, name_finish_folder, documents,
                                           now_doc, all_doc, current_progress, percent, logging, event, window_check,
                                           line_doing, line_progress, progress_value, info_value)
        if answer['status'] == 'error':
            return {'status': 'error', 'text': answer['text'], 'trace': answer['trace']}
        if answer['status'] == 'cancel':
            return {'status': 'cancel', 'text': '', 'trace': ''}
        if answer['status'] == 'warning':
            return {'status': 'warning', 'text': answer['text'], 'trace': ''}
        return {'status': 'warning' if errors else 'success', 'text': errors, 'trace': ''}
    except BaseException as error:
        return {'status': 'error', 'text': error, 'trace': traceback.format_exc()}
