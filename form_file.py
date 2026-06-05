import os
import traceback
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from pathlib import Path

from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def form_file(incoming_data: dict, current_progress: float, now_doc: int, all_doc: int, line_doing, line_progress,
              progress_value, event, window_check, info_value) -> dict:
    logging = incoming_data['logging']
    try:
        save_file_name = f"{Path(incoming_data['unformat_file']).stem}" \
                         f"_sort{Path(incoming_data['unformat_file']).suffix}"
        # Получаем ширину столбцов
        logging.info('Читаем файл')
        line_doing.emit(f"Читаем файл {Path(incoming_data['unformat_file']).name}")
        wb = load_workbook(incoming_data['unformat_file'])
        ws = wb.active
        column_sheet = [ws.column_dimensions[get_column_letter(col[0].column)].width for col in ws.iter_cols()]
        row_height = ws.row_dimensions[1].height
        df = pd.read_excel(incoming_data['unformat_file'], header=None)
        nan_col = df.columns[df.isna().all()].tolist()
        rows, cols = np.where(df.isna())
        errors = [f"Пропущено значение в ячейке {row + 1}{get_column_letter(col + 1)}"
                  for row, col in zip(rows, cols) if col not in nan_col]
        col_for_copy = [col for col in cols if col not in nan_col]
        for index, column in enumerate(col_for_copy):
            for col in df.columns:
                if not set(df[column]).isdisjoint(df[col]):
                    errors[index] = f"{errors[index]} (замена из колонки {get_column_letter(col + 1)})"
                    break
        current_progress += 25
        line_progress.emit(f'Выполнено {int(current_progress)} %')
        progress_value.emit(int(current_progress))
        line_doing.emit(f"Ищем уникальные колонки")
        column_name = df.iloc[0]
        df = df.iloc[1:]
        column_list = [df[column].to_numpy().tolist() for column in df.columns]
        uniq_column = {}
        columns = []
        for index, column in enumerate(column_list):
            str_list = [str(col) for col in column]
            if len(set(str_list)) == 1 and str_list[0] == 'nan' and index > 3:
                continue
            string_col = ','.join(str_list)
            if string_col not in columns:
                columns.append(string_col)
                uniq_column[index] = [column_name[index], *column]
        df = pd.DataFrame(uniq_column)
        similar_pairs = []
        df_similar = df[1:]
        cols = df_similar.columns
        for i in range(len(cols)):
            for j in range(i + 1, len(cols)):
                matches = (df_similar[cols[i]] == df_similar[cols[j]])
                match_pct = (matches.sum() / len(df_similar)) * 100
                if match_pct >= 70:
                    mismatches_df = df_similar[~matches][[cols[i], cols[j]]].copy()
                    similar_pairs.append({
                        'col1': df.iloc[0, i],
                        'col2': df.iloc[0, j],
                        'rows': mismatches_df.index.to_list(),
                        'match_percent': match_pct
                    })
        current_progress += 50
        line_progress.emit(f'Выполнено {int(current_progress)} %')
        progress_value.emit(int(current_progress))
        line_doing.emit(f"Заполняем новый файл {save_file_name}")
        save_file = Path(Path(incoming_data['unformat_file']).parent, save_file_name)
        df.dropna(axis=1, inplace=True)
        column_width = max([column_sheet[index] for index in list(uniq_column.keys())])
        writer = pd.ExcelWriter(save_file)
        df.to_excel(writer, sheet_name='sort', index=False, header=False, na_rep='NaN')
        for column in df:
            col_idx = df.columns.get_loc(column)
            writer.sheets['sort'].set_column(col_idx, col_idx, column_width)
        writer.sheets['sort'].set_row(0, row_height)
        workbook = writer.book
        worksheet = writer.sheets['sort']
        format_row = workbook.add_format({'text_wrap': True})
        worksheet.set_row(0, None, format_row)
        writer.close()
        wb = load_workbook(str(save_file))
        ws = wb.active
        write_xlsx = False
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type='solid')
        for similar in similar_pairs:
            write_xlsx = True
            col1 = col2 = 0
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value == similar['col1']:
                    col1 = col
                if cell_value == similar['col2']:
                    col2 = col
                if col1 != 0 and col2 != 0:
                    break
            for cell in ws[get_column_letter(col1)]:
                cell.fill = yellow_fill
            for cell in ws[get_column_letter(col2)]:
                cell.fill = yellow_fill
            for row in similar['rows']:
                cell1 = f"{get_column_letter(col1)}{row + 1}"
                cell2 = f"{get_column_letter(col2)}{row + 1}"
                ws[cell1].fill = orange_fill
                ws[cell2].fill = orange_fill
        if write_xlsx:
            while True:
                try:
                    wb.save(str(save_file))
                    break
                except PermissionError:
                    info_value.emit('Вопрос?',
                                    f"Файл {save_file.name} открыт. Для сохранения ошибок в файл закройте "
                                    f"его и нажмите «Да». Если нажать «Нет» ошибки не будут выделены!",
                                    "")
                    event.clear()
                    event.wait()
                if window_check.answer is False:
                    break
        current_progress += 25
        line_progress.emit(f'Выполнено {int(current_progress)} %')
        progress_value.emit(int(current_progress))
        os.startfile(save_file)
        return {'status': 'warning' if errors else 'success', 'text': errors, 'trace': ''}
    except BaseException as ex:
        return {'status': 'error', 'text': ex, 'trace': traceback.format_exc()}
